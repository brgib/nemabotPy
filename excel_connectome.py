"""excel_connectome.py

Load a connectome from an Excel adjacency matrix (.xlsx) and expose a fast
"apply" method, similar in spirit to connectome.py's generated per-neuron
functions.

Workbook expectation (works with the provided "SI 7 ... adjacency matrices"):
- A sheet contains an adjacency matrix.
- Column headers = postsynaptic neuron names (receivers)
- Row headers    = presynaptic neuron names (senders)
- Cell value     = weight to add when the presynaptic neuron fires

This module:
- Detects the header row/column via simple heuristics
- Builds: edges[presyn] = [(postsyn, weight), ...]
- Can populate/refresh the global postsynaptic dict used by Nemabot.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Any, Iterable

import openpyxl


Weight = float
EdgeList = List[Tuple[str, Weight]]


@dataclass
class ExcelConnectome:
    sheet_name: str
    neurons: List[str]                 # union of row+col labels (unique, ordered)
    edges: Dict[str, EdgeList]         # presynaptic -> list[(postsyn, weight)]

    # ---------------------------- heuristics ---------------------------------

    @staticmethod
    def _is_label(v: Any) -> bool:
        return isinstance(v, str) and v.strip() != ""

    @staticmethod
    def _find_header_row(ws) -> int:
        """Row that most looks like a label row (many non-empty strings)."""
        best_row, best_score = 1, -1
        max_r = min(ws.max_row, 60)
        max_c = min(ws.max_column, 260)
        for r in range(1, max_r + 1):
            score = 0
            for c in range(1, max_c + 1):
                if ExcelConnectome._is_label(ws.cell(r, c).value):
                    score += 1
            if score > best_score:
                best_row, best_score = r, score
        return best_row

    @staticmethod
    def _find_header_col(ws) -> int:
        """Column that most looks like a label column (many non-empty strings)."""
        best_col, best_score = 1, -1
        max_r = min(ws.max_row, 420)
        max_c = min(ws.max_column, 60)
        for c in range(1, max_c + 1):
            score = 0
            for r in range(1, max_r + 1):
                if ExcelConnectome._is_label(ws.cell(r, c).value):
                    score += 1
            if score > best_score:
                best_col, best_score = c, score
        return best_col

    @staticmethod
    def _collect_labels_in_row(ws, header_row: int, start_col: int) -> List[str]:
        labels: List[str] = []
        blank_run = 0
        for c in range(start_col, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v is None:
                blank_run += 1
                if blank_run >= 5:
                    break
                continue
            blank_run = 0
            if ExcelConnectome._is_label(v):
                labels.append(str(v).strip())
        return labels

    @staticmethod
    def _collect_labels_in_col(ws, header_col: int, start_row: int) -> List[str]:
        labels: List[str] = []
        blank_run = 0
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(r, header_col).value
            if v is None:
                blank_run += 1
                if blank_run >= 5:
                    break
                continue
            blank_run = 0
            if ExcelConnectome._is_label(v):
                labels.append(str(v).strip())
        return labels

    @classmethod
    def load(cls, xlsx_path: str, sheet_name: str) -> "ExcelConnectome":
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        hr = cls._find_header_row(ws)
        hc = cls._find_header_col(ws)

        # We assume [hr, hc] is the corner cell (blank or label like "Neuron")
        col_labels = cls._collect_labels_in_row(ws, hr, hc + 1)   # postsynaptic
        row_labels = cls._collect_labels_in_col(ws, hc, hr + 1)   # presynaptic

        if not col_labels or not row_labels:
            raise ValueError(
                f"Could not detect labels in '{sheet_name}'. "
                f"Detected header_row={hr}, header_col={hc}."
            )

        # Build edges
        edges: Dict[str, EdgeList] = {rn: [] for rn in row_labels}
        max_cols = min(len(col_labels), ws.max_column - (hc + 0))
        max_rows = min(len(row_labels), ws.max_row - (hr + 0))

        for i, presyn in enumerate(row_labels[:max_rows]):
            r = hr + 1 + i
            out: EdgeList = []
            for j, postsyn in enumerate(col_labels[:max_cols]):
                c = hc + 1 + j
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if isinstance(v, bool):
                    continue
                try:
                    w = float(v)
                except Exception:
                    continue
                if w == 0.0:
                    continue
                out.append((postsyn, w))
            edges[presyn] = out

        # Neuron set = union of row+col labels, keeping order: rows then missing cols
        seen = set()
        neurons: List[str] = []
        for n in row_labels + col_labels:
            if n not in seen:
                seen.add(n)
                neurons.append(n)

        return cls(sheet_name=sheet_name, neurons=neurons, edges=edges)

    # ------------------------------ runtime ----------------------------------

    def apply(self, presyn: str, postsynaptic: Dict[str, list], next_state_index: int) -> None:
        """Apply outgoing edges of 'presyn' to postsynaptic[*][next_state_index]."""
        out = self.edges.get(presyn)
        if not out:
            return
        for target, w in out:
            if target in postsynaptic:
                postsynaptic[target][next_state_index] += w

    def populate_postsynaptic(
        self,
        postsynaptic: Dict[str, list],
        extra_names: Optional[Iterable[str]] = None,
        init_vector: Optional[List[float]] = None,
    ) -> None:
        """(Re)initialize the global postsynaptic dict with neurons (+ optional extras).

        Nemabot expects postsynaptic[name] = [0,0,0,0,0] at minimum.
        """
        if init_vector is None:
            init_vector = [0, 0, 0, 0, 0]
        postsynaptic.clear()

        names = list(self.neurons)
        if extra_names:
            for n in extra_names:
                if n not in postsynaptic and n not in names:
                    names.append(n)

        for n in names:
            postsynaptic[n] = list(init_vector)


def list_workbook_sheets(xlsx_path: str) -> List[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    return list(wb.sheetnames)
